import gradio as gr
import zipfile
import tempfile
import os
import pandas as pd
from pathlib import Path
import re
from openpyxl import load_workbook
import shutil
import json
from typing import List, Dict, Tuple
import time
import traceback

# ==================== ФУНКЦИИ ОБРАБОТКИ ДАННЫХ ====================

def extract_attributes_from_template(template_file_path: str) -> List[str]:
    """
    Извлекает атрибуты из файла шаблона (.xlsx)
    Атрибуты находятся во 2 или 3 строке, начиная с 'Объект данных'
    и заканчивая 'Базовая единица измерения' (не включая)
    """
    try:
        wb = load_workbook(template_file_path, data_only=True, read_only=True)
        ws = wb.active
        
        attributes = []
        found_start = False
        stop_attributes = False
        
        # Ищем строку с атрибутами (2 или 3 строка)
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if stop_attributes:
                break
                
            row_values = [str(cell) if cell is not None else '' for cell in row]
            row_str = ' '.join(row_values)
            
            if 'Объект данных' in row_str and not found_start:
                found_start = True
                for cell in row:
                    if stop_attributes:
                        break
                    if cell:
                        cell_str = str(cell).strip()
                        if 'Базовая единица измерения' in cell_str:
                            stop_attributes = True
                            break
                        if cell_str and 'Объект данных' in cell_str:
                            attributes.append('Объект данных')
                        elif cell_str and cell_str not in attributes:
                            attributes.append(cell_str)
    
    except Exception as e:
        print(f"Ошибка чтения шаблона: {e}")
        attributes = []
    
    # Исключаем нежелательные атрибуты
    excluded_attributes = [
        'Код из системы источника',
        'Наименование',
        'Наименование из системы источника',
        'Полное наименование',
        'Статус'
    ]
    
    # Фильтрация атрибутов
    filtered_attributes = []
    for attr in attributes:
        if attr and str(attr).strip():
            attr_lower = str(attr).strip().lower()
            is_excluded = False
            for excluded in excluded_attributes:
                excluded_lower = excluded.lower()
                if (excluded_lower == attr_lower or 
                    excluded_lower in attr_lower or 
                    attr_lower in excluded_lower):
                    is_excluded = True
                    break
            
            if not is_excluded:
                filtered_attributes.append(str(attr).strip())
    
    return filtered_attributes

def extract_columns_from_records(records_file_path: str) -> List[str]:
    """Извлекает точные названия столбцов из файла записей"""
    try:
        # Читаем только заголовки
        df = pd.read_excel(records_file_path, nrows=0)
        columns = [str(col).strip() for col in df.columns.tolist()]
        return columns
    except Exception as e:
        print(f"Ошибка чтения заголовков записей: {e}")
        return []

def match_attributes(template_attributes: List[str], record_columns: List[str]) -> Tuple[List[str], List[str]]:
    """
    Сопоставляет атрибуты по 100% совпадению
    Возвращает: (совпавшие_атрибуты, отсутствующие_атрибуты)
    """
    matched = []
    missing = []
    
    for attr in template_attributes:
        attr_clean = attr.strip()
        if attr_clean in record_columns:
            matched.append(attr_clean)
        else:
            missing.append(attr_clean)
    
    return matched, missing

def extract_values_from_records(records_file_path: str, matched_attributes: List[str]) -> List[Dict[str, str]]:
    """Извлекает значения из файла записей для сопоставленных атрибутов"""
    values = []
    
    try:
        df = pd.read_excel(records_file_path)
        df.columns = [str(col).strip() for col in df.columns]
        
        for _, row in df.iterrows():
            record_values = {}
            for attr in matched_attributes:
                if attr in df.columns and pd.notna(row[attr]):
                    value = str(row[attr]).strip()
                    if value and value.lower() not in ['', 'nan', 'none', 'null']:
                        record_values[attr] = value
            
            if record_values:  # Добавляем только если есть значения
                values.append(record_values)
                
    except Exception as e:
        print(f"Ошибка извлечения значений: {e}")
    
    return values

def process_extracted_structure(extracted_path: Path, progress_callback=None) -> Tuple[List[Dict], List[Dict]]:
    """
    Обрабатывает распакованную структуру папок
    Возвращает: (данные, статистика_отсутствия)
    """
    all_data = []
    missing_stats = []
    
    # Ищем корневую папку 'Онтология ГРМ' или берем первую папку
    root_path = None
    for item in extracted_path.rglob('Онтология ГРМ'):
        if item.is_dir():
            root_path = item
            break
    
    if not root_path:
        # Если не нашли, берем первую папку в архиве
        items = list(extracted_path.iterdir())
        if items and items[0].is_dir():
            root_path = items[0]
    
    if not root_path:
        raise ValueError("В архиве не найдена корректная структура папок")
    
    print(f"Корневая папка для обработки: {root_path}")
    
    # Собираем все папки для обработки
    all_dirs = []
    for class_dir in root_path.iterdir():
        if class_dir.is_dir():
            for template_dir in class_dir.iterdir():
                if template_dir.is_dir():
                    all_dirs.append((class_dir.name, template_dir.name, template_dir))
    
    total_dirs = len(all_dirs)
    
    # Обрабатываем каждую папку
    for idx, (class_name, template_name, template_dir) in enumerate(all_dirs):
        if progress_callback:
            progress = (idx + 1) / total_dirs * 100
            progress_callback(progress, f"Обработка: {class_name}/{template_name}")
        
        # Ищем файлы в папке шаблона
        template_file = None
        records_file = None
        
        for file in template_dir.iterdir():
            if file.is_file():
                filename = file.name.lower()
                if 'шаблон.xlsx' in filename:
                    template_file = file
                elif 'предзап.xlsx' in filename:
                    records_file = file
        
        if template_file and records_file:
            # Извлекаем атрибуты
            template_attributes = extract_attributes_from_template(str(template_file))
            
            if template_attributes:
                # Извлекаем столбцы
                record_columns = extract_columns_from_records(str(records_file))
                
                # Сопоставляем
                matched_attrs, missing_attrs = match_attributes(template_attributes, record_columns)
                
                # Добавляем в статистику отсутствия
                for missing_attr in missing_attrs:
                    missing_stats.append({
                        'Класс': class_name,
                        'Шаблон': template_name,
                        'Атрибут': missing_attr,
                        'Статус': 'Отсутствует в файле записей'
                    })
                
                # Извлекаем значения
                if matched_attrs:
                    values = extract_values_from_records(str(records_file), matched_attrs)
                    
                    # Формируем записи
                    for record in values:
                        for attr, value in record.items():
                            all_data.append({
                                'Атрибут': attr,
                                'Шаблон': template_name,
                                'Класс': class_name,
                                'Значение': value
                            })
    
    return all_data, missing_stats

def collapse_duplicate_values(data: List[Dict]) -> Dict[str, set]:
    """Схлопывает одинаковые значения в одном атрибуте"""
    collapsed = {}
    
    for item in data:
        attr = item['Атрибут']
        value = item['Значение']
        
        if attr not in collapsed:
            collapsed[attr] = set()
        
        if value:
            collapsed[attr].add(value)
    
    return collapsed

def create_output_files(data: List[Dict], missing_stats: List[Dict], output_dir: Path) -> Dict[str, str]:
    """
    Создает все выходные файлы
    Возвращает словарь с путями к созданным файлам
    """
    created_files = {}
    
    # 1. Создаем отдельные файлы для каждого атрибута
    if data:
        collapsed_data = collapse_duplicate_values(data)
        
        # Папка для вспомогательных справочников
        справочники_dir = output_dir / "Вспомогательные справочники"
        справочники_dir.mkdir(parents=True, exist_ok=True)
        
        for attr, values in collapsed_data.items():
            if values:  # Если есть значения
                # Очищаем имя файла
                clean_attr_name = re.sub(r'[<>:"/\\|?*]', '_', attr)
                filename = f"{clean_attr_name}.xlsx"
                filepath = справочники_dir / filename
                
                # Создаем DataFrame
                df = pd.DataFrame({
                    'Атрибут': [attr] * len(values),
                    'Значение': sorted(values)
                })
                
                # Сохраняем
                df.to_excel(filepath, index=False)
                created_files[f"справочники/{filename}"] = str(filepath)
        
        # 2. Создаем сводный файл
        summary_file = output_dir / "Сводные_данные.xlsx"
        df_all = pd.DataFrame(data)
        df_all.to_excel(summary_file, index=False)
        created_files["сводный.xlsx"] = str(summary_file)
    
    # 3. Создаем файл статистики
    if missing_stats:
        stats_file = output_dir / "Статистика_отсутствующих.xlsx"
        df_missing = pd.DataFrame(missing_stats)
        df_missing.to_excel(stats_file, index=False)
        created_files["статистика.xlsx"] = str(stats_file)
    
    # 4. Создаем README с результатами
    readme_file = output_dir / "README.txt"
    with open(readme_file, 'w', encoding='utf-8') as f:
        f.write("РЕЗУЛЬТАТЫ ОБРАБОТКИ ОНТОЛОГИИ ГРМ\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"Дата обработки: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        if data:
            f.write(f"Обработано записей: {len(data):,}\n")
            f.write(f"Уникальных атрибутов: {len(collapsed_data):,}\n")
        
        if missing_stats:
            f.write(f"Найдено отсутствующих атрибутов: {len(missing_stats):,}\n")
        
        f.write("\nСозданные файлы:\n")
        for filename in created_files.keys():
            f.write(f"• {filename}\n")
    
    created_files["readme.txt"] = str(readme_file)
    
    return created_files

def create_results_zip(output_dir: Path) -> str:
    """Создает ZIP-архив со всеми результатами"""
    zip_path = output_dir.parent / "результаты_обработки.zip"
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, output_dir.parent)
                zipf.write(file_path, arcname)
    
    return str(zip_path)

# ==================== GRADO ИНТЕРФЕЙС ====================

def update_progress(progress, message, progress_bar, status_text):
    """Обновляет прогресс-бар и статус"""
    if progress_bar is not None:
        progress_bar(progress / 100, desc=message)
    if status_text is not None:
        return status_text + f"\n{message}"
    return ""

def process_ontology_archive(zip_file_path, progress=gr.Progress()):
    """
    Основная функция обработки архива
    Возвращает: (путь_к_результатам, отчет, файлы_для_скачивания)
    """
    if not zip_file_path:
        return None, "❌ Пожалуйста, загрузите ZIP-архив", []
    
    # Создаем временные директории
    temp_dir = tempfile.mkdtemp(prefix="ontology_")
    extracted_dir = Path(temp_dir) / "extracted"
    output_dir = Path(temp_dir) / "results"
    
    try:
        # Шаг 1: Распаковка архива
        progress(0, desc="📦 Распаковка архива...")
        time.sleep(0.5)
        
        extracted_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                zip_ref.extractall(extracted_dir)
        except zipfile.BadZipFile:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None, "❌ Ошибка: поврежденный или некорректный ZIP-архив", []
        
        # Проверяем, что что-то распаковалось
        extracted_items = list(extracted_dir.iterdir())
        if not extracted_items:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None, "❌ Архив пуст или не содержит данных", []
        
        progress(0.2, desc="✅ Архив распакован")
        
        # Шаг 2: Обработка данных
        progress(0.2, desc="🔍 Анализ структуры данных...")
        
        # Функция для обновления прогресса
        def update_progress_callback(pct, msg):
            progress(0.2 + pct/100 * 0.6, desc=msg)
        
        try:
            all_data, missing_stats = process_extracted_structure(
                extracted_dir, 
                update_progress_callback
            )
        except Exception as e:
            error_details = traceback.format_exc()
            print(f"Ошибка обработки: {error_details}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None, f"❌ Ошибка обработки данных: {str(e)}", []
        
        if not all_data and not missing_stats:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None, "⚠️ В архиве не найдены данные для обработки", []
        
        progress(0.8, desc="✅ Данные обработаны")
        
        # Шаг 3: Создание выходных файлов
        progress(0.8, desc="💾 Сохранение результатов...")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        created_files = create_output_files(all_data, missing_stats, output_dir)
        
        progress(0.9, desc="📦 Упаковка результатов...")
        
        # Создаем ZIP-архив
        result_zip = create_results_zip(output_dir)
        
        progress(1.0, desc="✅ Готово!")
        
        # Формируем отчет
        report_lines = [
            "=" * 50,
            "📊 ОТЧЕТ ОБ ОБРАБОТКЕ",
            "=" * 50,
            f"📅 Дата: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "✅ РЕЗУЛЬТАТЫ:"
        ]
        
        if all_data:
            df_all = pd.DataFrame(all_data)
            unique_attrs = df_all['Атрибут'].nunique()
            unique_values = df_all['Значение'].nunique()
            unique_templates = df_all['Шаблон'].nunique()
            unique_classes = df_all['Класс'].nunique()
            
            report_lines.extend([
                f"• Обработано записей: {len(all_data):,}",
                f"• Уникальных атрибутов: {unique_attrs:,}",
                f"• Уникальных значений: {unique_values:,}",
                f"• Шаблонов: {unique_templates:,}",
                f"• Классов: {unique_classes:,}"
            ])
        
        if missing_stats:
            df_missing = pd.DataFrame(missing_stats)
            report_lines.extend([
                "",
                "⚠️ ПРОБЛЕМЫ СОПОСТАВЛЕНИЯ:",
                f"• Атрибутов в шаблонах, отсутствующих в данных: {len(missing_stats):,}",
                f"• Уникальных отсутствующих атрибутов: {df_missing['Атрибут'].nunique():,}",
                f"• Шаблонов с проблемами: {df_missing['Шаблон'].nunique():,}"
            ])
        
        report_lines.extend([
            "",
            "📁 СОЗДАННЫЕ ФАЙЛЫ:",
            "• результаты_обработки.zip - архив со всеми файлами"
        ])
        
        for category in ["справочники/", "сводный.xlsx", "статистика.xlsx", "readme.txt"]:
            if any(k.startswith(category) for k in created_files.keys()):
                display_name = {
                    "справочники/": "Вспомогательные справочники (отдельные файлы по атрибутам)",
                    "сводный.xlsx": "Сводные_данные.xlsx - все извлеченные записи",
                    "статистика.xlsx": "Статистика_отсутствующих.xlsx - проблемы сопоставления",
                    "readme.txt": "README.txt - описание результатов"
                }.get(category, category)
                report_lines.append(f"  ◦ {display_name}")
        
        report_lines.extend([
            "",
            "=" * 50,
            "✅ ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО!",
            "=" * 50
        ])
        
        report = "\n".join(report_lines)
        
        # Создаем список файлов для отображения в интерфейсе
        display_files = [
            (result_zip, "результаты_обработки.zip")
        ]
        
        return result_zip, report, display_files
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Критическая ошибка: {error_details}")
        
        # Пытаемся очистить временные файлы
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass
        
        return None, f"❌ Критическая ошибка: {str(e)}\n\nДетали: {error_details[:500]}", []

# ==================== СОЗДАНИЕ ИНТЕРФЕЙСА ====================

with gr.Blocks(
    theme=gr.themes.Soft(
        primary_hue="blue",
        secondary_hue="purple",
    ),
    title="🏭 Обработчик онтологии ГРМ",
    css="""
    .gradio-container { max-width: 1200px !important; }
    .success-box { background-color: #e8f5e9; padding: 20px; border-radius: 10px; border-left: 5px solid #4caf50; }
    .warning-box { background-color: #fff3e0; padding: 20px; border-radius: 10px; border-left: 5px solid #ff9800; }
    .file-info { font-family: monospace; background-color: #f5f5f5; padding: 10px; border-radius: 5px; }
    """
) as demo:
    
    # Заголовок
    gr.HTML("""
    <div style="text-align: center;">
        <h1 style="color: #1a237e;">🏭 Обработчик онтологии ГРМ</h1>
        <p style="color: #5c6bc0; font-size: 1.1em;">
            Веб-интерфейс для обработки структурированных данных онтологии
        </p>
    </div>
    """)
    
    # Основная информация
    with gr.Row():
        with gr.Column(scale=2):
            gr.Markdown("""
            ### 📋 Как использовать:
            
            1. **Подготовьте архив**: Упакуйте папку с онтологией в **ZIP-архив**
            2. **Загрузите архив**: Используйте форму справа
            3. **Обработайте**: Нажмите кнопку "Запустить обработку"
            4. **Скачайте результаты**: Получите архив с обработанными данными
            
            ### 📁 Требуемая структура в архиве:
            ```
            Ваш_архив.zip/
            ├── Онтология ГРМ/                    # Корневая папка
            │   ├── Класс1/                       # Папка класса
            │   │   ├── Шаблон1/                  # Папка шаблона
            │   │   │   ├── *Шаблон.xlsx          # Файл шаблона
            │   │   │   └── *ПредЗап.xlsx         # Файл записей
            │   │   └── Шаблон2/
            │   └── Класс2/
            ```
            
            *Файлы могут иметь другие названия, но должны содержать ключевые слова*
            """)
            
            # Информационная панель
            with gr.Accordion("📖 Подробная инструкция", open=False):
                gr.Markdown("""
                ### 🔍 Что делает обработчик:
                
                1. **Извлекает атрибуты** из файлов шаблонов (строки 2-3, от "Объект данных" до "Базовая единица измерения")
                2. **Исключает системные атрибуты**: "Код из системы источника", "Наименование", и т.д.
                3. **Сопоставляет атрибуты** по 100% совпадению с файлами записей
                4. **Извлекает значения** и схлопывает дубликаты
                5. **Создает результаты**:
                   - Отдельные файлы для каждого атрибута
                   - Сводный файл со всеми данными
                   - Статистику по отсутствующим атрибутам
                
                ### ⚠️ Важные замечания:
                - Максимальный размер архива: **1 ГБ** (ограничение Hugging Face Spaces)
                - Время обработки зависит от объема данных
                - Все файлы обрабатываются в памяти, большие архивы могут вызвать ошибки
                
                ### 🛠️ Технические детали:
                - Формат файлов: Excel (.xlsx)
                - Кодировка: UTF-8
                - Поддерживаемые библиотеки: pandas, openpyxl
                """)
        
        with gr.Column(scale=1):
            # Панель загрузки
            gr.Markdown("### 📤 Загрузка архива")
            
            zip_input = gr.File(
                label="Выберите ZIP-архив",
                file_types=[".zip"],
                type="filepath",
                height=100
            )
            
            with gr.Row():
                process_btn = gr.Button(
                    "🚀 Запустить обработку",
                    variant="primary",
                    scale=2
                )
                
                clear_btn = gr.Button(
                    "🔄 Очистить",
                    variant="secondary",
                    scale=1
                )
            
            # Примеры (если есть тестовые данные)
            try:
                import os
                if os.path.exists("пример_архива.zip"):
                    gr.Examples(
                        examples=[["пример_архива.zip"]],
                        inputs=[zip_input],
                        label="📁 Пример архива для тестирования"
                    )
            except:
                pass
    
    # Разделитель
    gr.Markdown("---")
    
    # Панель прогресса и результатов
    with gr.Row():
        with gr.Column(scale=3):
            # Отчет
            report_output = gr.Textbox(
                label="📋 Отчет о работе",
                lines=15,
                interactive=False,
                show_copy_button=True
            )
            
            # Прогресс-бар (невидимый, используется для обновлений)
            progress_bar = gr.Progress(visible=False)
        
        with gr.Column(scale=1):
            # Результаты для скачивания
            gr.Markdown("### 📥 Результаты")
            
            result_files = gr.Files(
                label="Скачать результаты",
                file_count="multiple",
                interactive=False,
                height=300
            )
            
            # Статистика обработки
            stats_display = gr.HTML("""
            <div style="background-color: #f5f5f5; padding: 15px; border-radius: 10px;">
                <h4 style="margin-top: 0;">📈 Ожидаемые результаты:</h4>
                <ul style="margin-bottom: 0;">
                    <li>Архив с обработанными данными</li>
                    <li>Отдельные файлы по атрибутам</li>
                    <li>Сводная статистика</li>
                    <li>Отчет об обработке</li>
                </ul>
            </div>
            """)
    
    # Обработчики событий
    def clear_all():
        """Очищает все поля"""
        return None, "", []
    
    # Привязка обработчиков
    process_btn.click(
        fn=process_ontology_archive,
        inputs=[zip_input],
        outputs=[gr.File(visible=False), report_output, result_files],
        show_progress="full"
    )
    
    clear_btn.click(
        fn=clear_all,
        inputs=[],
        outputs=[zip_input, report_output, result_files]
    )
    
    # Футер
    gr.HTML("""
    <div style="text-align: center; margin-top: 30px; padding: 20px; background-color: #f5f5f5; border-radius: 10px;">
        <p style="color: #666; font-size: 0.9em;">
            🏭 Обработчик онтологии ГРМ | 
            📧 Поддержка: ваша_почта@example.com |
            ⚠️ Для больших архивов рекомендуем использовать локальную версию
        </p>
        <p style="color: #999; font-size: 0.8em; margin-top: 10px;">
            Версия 1.0.0 | Обработка выполняется на сервере Hugging Face Spaces
        </p>
    </div>
    """)

# ==================== ЗАПУСК ПРИЛОЖЕНИЯ ====================

if __name__ == "__main__":
    # Для локального тестирования
    demo.launch(
        server_name="0.0.0.0",
        server_port=7860,
        share=False,
        show_error=True
    )